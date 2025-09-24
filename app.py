# app.py — Clustering + Batch Navigation + Skip/Mark + Clickable Image Popups
# with Map-Click Start & Address Search (no GPS/IP needed)
# ----------------------------------------------------------------------------
# Added: JSON persistence for visited/skipped + photo counts (no photos saved to disk!)
# - Progress JSON: ./progress/<dataset_id>.json
# - Remembers visited/skipped and how many photos uploaded per Issue ID
# - Actual photos remain session-only (in memory); use the ZIP download to keep them

import math
import io
import json
import hashlib
from pathlib import Path
from datetime import datetime
import tempfile
import requests
import numpy as np
import pandas as pd
import streamlit as st

# Optional heavy deps — guarded (GeoPandas/Fiona may be unavailable on Streamlit Cloud)
try:
    import geopandas as gpd
    HAVE_GPD = True
except Exception:
    HAVE_GPD = False

try:
    import fiona
    HAS_FIONA = True
except Exception:
    HAS_FIONA = False

from sklearn.cluster import DBSCAN
import folium
from folium.plugins import MarkerCluster
from openpyxl import load_workbook
from streamlit_folium import st_folium  # << for map click
from html import escape

# ----------------- JSON Progress Persistence -----------------
PROGRESS_ROOT = Path("./progress")
PROGRESS_ROOT.mkdir(parents=True, exist_ok=True)

def dataset_fingerprint(file_bytes: bytes) -> str:
    """Stable ID for the uploaded CSV so we can restore progress on re-uploads."""
    return hashlib.sha1(file_bytes).hexdigest()

def progress_path(dataset_id: str) -> Path:
    return PROGRESS_ROOT / f"{dataset_id}.json"

def load_progress(dataset_id: str):
    """Return (visited_set, skipped_set, photo_counts_dict)."""
    p = progress_path(dataset_id)
    if p.exists():
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
        visited = set(map(str, data.get("visited_ticket_ids", [])))
        skipped = set(map(str, data.get("skipped_ticket_ids", [])))
        photo_counts = data.get("uploaded_after_photos", {})  # {issue_id: count}
        return visited, skipped, photo_counts
    return set(), set(), {}

def save_progress(dataset_id: str, visited_ids: set, skipped_ids: set, photo_counts: dict):
    """Persist visited/skipped and photo counts (no image bytes/paths)."""
    if not dataset_id:
        return
    p = progress_path(dataset_id)
    payload = {
        "visited_ticket_ids": sorted(list(visited_ids)),
        "skipped_ticket_ids": sorted(list(skipped_ids)),
        "uploaded_after_photos": photo_counts,
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    with p.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

# ----------------- Constants -----------------
EARTH_RADIUS_M = 6_371_000.0
REQUIRED_COLS = {
    'ISSUE ID', 'CITY', 'ZONE', 'WARD', 'SUBCATEGORY', 'CREATED AT',
    'STATUS', 'LATITUDE', 'LONGITUDE', 'BEFORE PHOTO', 'AFTER PHOTO', 'ADDRESS'
}

# ----------------- Helpers -----------------
def normalize_subcategory(series: pd.Series):
    return series.astype(str).str.strip().str.lower()

def haversine_m(lat1, lon1, lat2, lon2):
    phi1, phi2 = math.radians(float(lat1)), math.radians(float(lat2))
    dphi = math.radians(float(lat2) - float(lat1))
    dlmb = math.radians(float(lon2) - float(lon1))
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlmb/2)**2
    return 2 * EARTH_RADIUS_M * math.asin(math.sqrt(a))

def google_maps_url(origin_lat, origin_lon, dest_lat, dest_lon, mode="driving", waypoints=None):
    base = "https://www.google.com/maps/dir/?api=1"
    parts = []
    # If origin is missing, omit it — Maps will start from live GPS on the device
    if origin_lat is not None and origin_lon is not None:
        parts.append(f"origin={origin_lat},{origin_lon}")
    parts.append(f"destination={dest_lat},{dest_lon}")
    mode = mode if mode in {"driving","walking","bicycling","transit"} else "driving"
    parts.append(f"travelmode={mode}")
    if waypoints:
        from urllib.parse import quote
        wp = "|".join([f"{lat},{lon}" for (lat,lon) in waypoints])
        parts.append(f"waypoints={quote(wp, safe='|,')}")
    return base + "&" + "&".join(parts)

def hyperlinkify_excel(excel_path):
    try:
        wb = load_workbook(excel_path)
        ws = wb["Clustering Application Summary"]
        for row in range(2, ws.max_row + 1):
            for col in (11, 12):  # BEFORE/AFTER PHOTO columns
                link = ws.cell(row, col).value
                if link and isinstance(link, str) and link.startswith(("http://", "https://")):
                    ws.cell(row, col).hyperlink = link
                    ws.cell(row, col).style = "Hyperlink"
        wb.save(excel_path)
    except Exception:
        pass

def load_wards_uploaded(file):
    """Ward file reader with KML guard. Returns GeoDataFrame or None."""
    if not HAVE_GPD:
        st.info("GeoPandas not available; ward overlay/join skipped.")
        return None
    try:
        suffix = file.name.split(".")[-1].lower()
        with tempfile.NamedTemporaryFile(delete=False, suffix="."+suffix) as tmp:
            tmp.write(file.read())
            tmp_path = tmp.name

        if suffix in ("geojson", "json"):
            gdf = gpd.read_file(tmp_path)
        elif suffix == "kml":
            if not HAS_FIONA:
                st.warning("KML requires Fiona/GDAL. Upload GeoJSON/JSON instead.")
                return None
            layers = fiona.listlayers(tmp_path)
            gdf = None
            for layer in layers:
                gdf_try = gpd.read_file(tmp_path, driver="KML", layer=layer)
                if not gdf_try.empty:
                    gdf = gdf_try
                    break
            if gdf is None:
                st.warning("No non-empty KML layers found.")
                return None
        else:
            st.warning("Unsupported ward file type. Use GeoJSON/JSON/KML.")
            return None

        if gdf.crs is None:
            gdf.set_crs(epsg=4326, inplace=True)
        elif gdf.crs.to_string() != "EPSG:4326":
            gdf = gdf.to_crs(epsg=4326)
        return gdf[gdf.geometry.notna() & ~gdf.geometry.is_empty]
    except Exception:
        return None

def geocode_query(q: str, timeout=6):
    """Address/place geocoder (OSM Nominatim). Returns (lat, lon) or (None, None)."""
    try:
        url = "https://nominatim.openstreetmap.org/search"
        params = {"q": q, "format": "json", "limit": 1}
        headers = {"User-Agent": "cluster-app/1.0"}
        r = requests.get(url, params=params, headers=headers, timeout=timeout)
        if r.status_code == 200:
            data = r.json()
            if data:
                return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception:
        pass
    return None, None

def build_nearest_neighbor_sequence(df_like: pd.DataFrame, start_lat: float, start_lon: float, limit: int):
    """Greedy NN sequence from a start; returns list of rows (dict-like)."""
    seq = []
    pool = df_like.copy()
    if pool.empty:
        return seq
    cur_lat, cur_lon = float(start_lat), float(start_lon)
    for _ in range(min(limit, len(pool))):
        pool['__dist'] = pool.apply(lambda r: haversine_m(cur_lat, cur_lon, r['LATITUDE'], r['LONGITUDE']), axis=1)
        nxt = pool.sort_values('__dist').iloc[0]
        seq.append(nxt)
        cur_lat, cur_lon = float(nxt['LATITUDE']), float(nxt['LONGITUDE'])
        pool = pool[pool['ISSUE ID'] != nxt['ISSUE ID']]
    return seq

def is_url(u):
    return isinstance(u, str) and u.startswith(("http://", "https://"))

# ----------------- App Setup -----------------
st.set_page_config(layout="wide", page_title="Clustering + Batch Navigation")
st.title("🗺️ Clustering + Batch Navigation (Map-Click Start • Skip/Mark • Downloads)")

# Session state
for k in ["visited_ticket_ids", "skipped_ticket_ids"]:
    if k not in st.session_state:
        st.session_state[k] = set()
if "batch_cursor" not in st.session_state:
    st.session_state.batch_cursor = 0  # pagination by 10
if "origin_lat" not in st.session_state:
    st.session_state.origin_lat = None
if "origin_lon" not in st.session_state:
    st.session_state.origin_lon = None
# Session-only photo storage (bytes, no disk)
if "uploaded_after_photos" not in st.session_state:
    # {issue_id: [ {bytes, saved_name, original_name, source, ward, status, ts_str}, ... ]}
    st.session_state.uploaded_after_photos = {}
# Restored counts from JSON
if "restored_photo_counts" not in st.session_state:
    st.session_state.restored_photo_counts = {}
# Dataset id in session
if "dataset_id" not in st.session_state:
    st.session_state.dataset_id = None

# ----------------- Inputs (unique keys) -----------------
csv_file = st.file_uploader("Upload CSV", type=["csv"], key="csv_uploader_main")
ward_file = st.file_uploader("Upload Wards file (optional)", type=["geojson","json","kml"], key="ward_uploader_main")

subcategory_option = st.selectbox(
    "Issue Subcategory",
    [
        "Pothole","Garbage dumped on public land","Unpaved Road","Broken Footpath / Divider",
        "Sand piled on roadsides + Mud/slit on roadside","Malba, bricks, bori, etc dumped on public land",
        "Construction/ demolition activity without safeguards","Encroachment-Building Materials Dumped on Road",
        "Burning of garbage, plastic, leaves, branches etc.","Overflowing Dustbins",
        "Barren land to be greened","Greening of Central Verges","Unsurfaced Parking Lots"
    ],
    key="sel_subcategory"
)
radius_m = st.number_input("Clustering radius (m)", 1, 1000, 15, key="num_radius")
min_samples = st.number_input("Minimum per cluster", 1, 100, 2, key="num_min_samples")

# ----------------- Main -----------------
if not csv_file:
    st.info("Upload the required CSV to proceed.")
    st.stop()

# Read raw bytes to compute dataset_id and parse CSV
csv_bytes = csv_file.getvalue()
dataset_id = dataset_fingerprint(csv_bytes)
st.session_state.dataset_id = dataset_id
st.caption(f"Dataset ID: `{dataset_id[:8]}…`")

# Load CSV from bytes (not direct file handle)
df = pd.read_csv(io.BytesIO(csv_bytes))
missing = list(REQUIRED_COLS - set(df.columns))
if missing:
    st.error(f"Missing required columns: {missing}")
    st.stop()

# Restore prior progress (visited/skipped/photo counts)
rest_visited, rest_skipped, rest_photo_counts = load_progress(dataset_id)
st.session_state.visited_ticket_ids |= rest_visited
st.session_state.skipped_ticket_ids |= rest_skipped
st.session_state.restored_photo_counts = rest_photo_counts

df['SUBCATEGORY_NORM'] = normalize_subcategory(df['SUBCATEGORY'])
df = df[df['SUBCATEGORY_NORM'] == subcategory_option.lower()].copy()

df['LATITUDE']  = pd.to_numeric(df['LATITUDE'], errors='coerce')
df['LONGITUDE'] = pd.to_numeric(df['LONGITUDE'], errors='coerce')
df = df[df['LATITUDE'].between(-90, 90, inclusive='both')]
df = df[df['LONGITUDE'].between(-180, 180, inclusive='both')]
df.dropna(subset=['LATITUDE','LONGITUDE'], inplace=True)

if df.empty:
    st.warning("No valid rows for the selected subcategory after cleaning LAT/LON.")
    st.stop()

# Clustering (haversine DBSCAN)
coords_deg = df[['LATITUDE','LONGITUDE']].to_numpy(dtype=float)
coords_rad = np.radians(coords_deg)
eps_rad = float(radius_m) / EARTH_RADIUS_M
db = DBSCAN(eps=eps_rad, min_samples=int(min_samples), metric='haversine', algorithm='ball_tree')
try:
    labels = db.fit_predict(coords_rad)
except ValueError as e:
    st.error(f"Clustering failed due to invalid input: {e}")
    st.stop()
df['CLUSTER NUMBER'] = labels
df['IS_CLUSTERED'] = df['CLUSTER NUMBER'] != -1

# Optional GeoPandas overlay/join
if HAVE_GPD:
    gdf_all = gpd.GeoDataFrame(df.copy(), geometry=gpd.points_from_xy(df['LONGITUDE'], df['LATITUDE']), crs="EPSG:4326")
    wards_gdf = load_wards_uploaded(ward_file) if ward_file else None
    if wards_gdf is not None and not wards_gdf.empty:
        try:
            gdf_all = gpd.sjoin(gdf_all, wards_gdf, how="left", predicate="within")
        except Exception:
            pass
else:
    gdf_all = df.copy()

# Excel summary for clusters
clustered = gdf_all[gdf_all['IS_CLUSTERED']]
excel_filename = "Clustering_Application_Summary.xlsx"
if not clustered.empty:
    (clustered.drop(columns=["geometry"]) if "geometry" in clustered.columns else clustered).to_excel(
        excel_filename, sheet_name='Clustering Application Summary', index=False
    )
    hyperlinkify_excel(excel_filename)

# ---------------- Start Point (Map-Click + Address Search + Manual) ----------------
st.subheader("Step 4: Set Start Point (No GPS Needed)")

# Address search (optional)
colA, colB = st.columns(2)
with colA:
    q = st.text_input("🔎 Search address / landmark (optional)", key="txt_addr_query")
    if st.button("Search & set start", key="btn_addr_geocode"):
        if q.strip():
            g_lat, g_lon = geocode_query(q.strip())
            if g_lat is not None:
                st.session_state.origin_lat = g_lat
                st.session_state.origin_lon = g_lon
                st.success(f"Start set from search: {g_lat:.6f}, {g_lon:.6f}")
                st.rerun()
            else:
                st.error("Could not geocode that query. Try a more specific address.")
with colB:
    # Manual coordinates (optional)
    man_lat = st.text_input("Manual latitude (optional)", key="txt_lat")
    man_lon = st.text_input("Manual longitude (optional)", key="txt_lon")
    if st.button("Set manual start", key="btn_set_manual"):
        try:
            st.session_state.origin_lat = float(man_lat)
            st.session_state.origin_lon = float(man_lon)
            st.success(f"Manual start set: {st.session_state.origin_lat:.6f}, {st.session_state.origin_lon:.6f}")
            st.rerun()
        except Exception:
            st.error("Invalid coordinates. Example: 26.8467, 80.9462")

# Map-click picker
st.caption("🗺️ Or click anywhere on the mini map to set your starting point:")
center_lat = float(df['LATITUDE'].mean())
center_lon = float(df['LONGITUDE'].mean())
mini = folium.Map(location=[center_lat, center_lon], zoom_start=12)
# Show existing start marker if set
if st.session_state.origin_lat is not None and st.session_state.origin_lon is not None:
    folium.Marker(
        [st.session_state.origin_lat, st.session_state.origin_lon],
        popup="Start here",
        icon=folium.Icon(color="green", icon="flag")
    ).add_to(mini)

click_state = st_folium(mini, height=320, width=None, key="mini_click", returned_objects=["last_clicked"])
if isinstance(click_state, dict):
    clicked = click_state.get("last_clicked")
    if clicked and "lat" in clicked and "lng" in clicked:
        st.session_state.origin_lat = float(clicked["lat"])
        st.session_state.origin_lon = float(clicked["lng"])
        st.success(f"Start set from map click: {st.session_state.origin_lat:.6f}, {st.session_state.origin_lon:.6f}")
        st.rerun()

origin_lat = st.session_state.origin_lat
origin_lon = st.session_state.origin_lon
if origin_lat is not None and origin_lon is not None:
    st.info(f"Using start: {origin_lat:.6f}, {origin_lon:.6f}")
else:
    st.warning("No start set. It’s okay—your Google Maps link will still start from the phone’s current GPS.")

# ---------------- Batch + Skip/Mark ----------------
st.subheader("Step 5: Batches of 10 — View • Skip • Mark")

# Filter out visited/skipped from the pool
pool = (gdf_all.drop(columns=["geometry"]) if "geometry" in getattr(gdf_all, "columns", []) else gdf_all).copy()
pool = pool[~pool['ISSUE ID'].astype(str).isin(st.session_state.visited_ticket_ids)]
pool = pool[~pool['ISSUE ID'].astype(str).isin(st.session_state.skipped_ticket_ids)]

st.caption(f"Remaining eligible tickets: {len(pool)} | Visited: {len(st.session_state.visited_ticket_ids)} | Skipped: {len(st.session_state.skipped_ticket_ids)}")

if pool.empty:
    st.info("No tickets remaining after filters/visited/skipped.")
    batch_df = pd.DataFrame(columns=['ISSUE ID','WARD','STATUS','LATITUDE','LONGITUDE','BEFORE PHOTO'])
else:
    # Seed NN: from start point; else from dataset centroid
    if origin_lat is None or origin_lon is None:
        seed_lat, seed_lon = float(pool['LATITUDE'].mean()), float(pool['LONGITUDE'].mean())
    else:
        seed_lat, seed_lon = origin_lat, origin_lon

    long_seq_rows = build_nearest_neighbor_sequence(pool, seed_lat, seed_lon, limit=min(200, len(pool)))
    seq_df = pd.DataFrame(long_seq_rows)

    if st.session_state.batch_cursor >= len(seq_df):
        st.session_state.batch_cursor = 0

    start = st.session_state.batch_cursor
    end = min(start + 10, len(seq_df))
    batch_df = seq_df.iloc[start:end].copy()

    if batch_df.empty and not seq_df.empty:
        st.session_state.batch_cursor = 0
        start, end = 0, min(10, len(seq_df))
        batch_df = seq_df.iloc[start:end].copy()

    # Show batch table (with Before Photo link)
    batch_df_display = batch_df[['ISSUE ID','WARD','STATUS','LATITUDE','LONGITUDE','BEFORE PHOTO']].reset_index(drop=True)
    batch_df_display.index = batch_df_display.index + 1  # 1-based row numbers
    # Linkify BEFORE PHOTO in the table
    def as_link_or_none(x):
        s = str(x or "").strip()
        return s if is_url(s) else None
    batch_df_display['BEFORE PHOTO'] = batch_df_display['BEFORE PHOTO'].apply(as_link_or_none)

    st.dataframe(
        batch_df_display,
        use_container_width=True,
        column_config={
            'LATITUDE': st.column_config.NumberColumn('LATITUDE', format="%.6f"),
            'LONGITUDE': st.column_config.NumberColumn('LONGITUDE', format="%.6f"),
            'BEFORE PHOTO': st.column_config.LinkColumn('Before Photo', display_text='Open'),
        }
    )

    # ------------- After Photo uploads panel (session-only, no disk) -------------
    st.markdown("### After Photo uploads for current batch")
    def _save_photo_bytes(img_bytes: bytes, issue_id: str, ward: str, status: str, original_name: str, source: str):
        ts_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        saved_name = f"{issue_id}_after_{ts_str}.jpg"
        st.session_state.uploaded_after_photos.setdefault(str(issue_id), []).append({
            "bytes": img_bytes,
            "saved_name": saved_name,
            "original_name": original_name,
            "source": source,               # "camera" or "upload"
            "ward": str(ward),
            "status": str(status),
            "ts_str": ts_str,
        })
        # Update JSON counts immediately
        counts = {iid: len(items) for iid, items in st.session_state.uploaded_after_photos.items()}
        save_progress(st.session_state.dataset_id,
                      st.session_state.visited_ticket_ids,
                      st.session_state.skipped_ticket_ids,
                      counts)

    for _, row in batch_df.iterrows():
        issue_id = str(row['ISSUE ID'])
        ward     = row.get('WARD', '')
        status   = row.get('STATUS', '')

        st.markdown(f"**Issue {issue_id}** — Ward {ward}, Status: {status}")
        cam = st.camera_input(f"Take photo ({issue_id})", key=f"cam_{issue_id}")
        if cam is not None:
            _save_photo_bytes(cam.getvalue(), issue_id, ward, status, original_name="camera.jpg", source="camera")
            st.success("Captured ✅")

        up = st.file_uploader(
            f"Upload photo ({issue_id})",
            type=["jpg","jpeg","png"],
            key=f"upl_{issue_id}"
        )
        if up is not None:
            _save_photo_bytes(up.read(), issue_id, ward, status, original_name=up.name, source="upload")
            st.success("Uploaded ✅")

        # Status line (session vs restored)
        if issue_id in st.session_state.uploaded_after_photos:
            cnt = len(st.session_state.uploaded_after_photos[issue_id])
            st.info(f"✅ {cnt} photo(s) saved in this session")
        elif issue_id in st.session_state.restored_photo_counts:
            cnt = st.session_state.restored_photo_counts[issue_id]
            st.info(f"✅ {cnt} photo(s) previously saved (no files in session)")
        else:
            st.warning("⚠️ No After Photo saved yet")

        st.divider()
    # ---------------------------------------------------------------------------

    # ---------------- ZIP Download of all session photos (with manifest) ----------------
    st.subheader("Download all clicked/uploaded After Photos")
    import zipfile, csv

    if st.session_state.uploaded_after_photos:
        # Build manifest rows and zip from memory (no disk writes)
        rows = []
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for issue_id, items in st.session_state.uploaded_after_photos.items():
                for it in items:
                    rel_path = f"after_photos/{issue_id}/{it['saved_name']}"
                    # write image bytes directly
                    zf.writestr(rel_path, it["bytes"])
                    rows.append([
                        issue_id,
                        it.get("ward",""),
                        it.get("status",""),
                        rel_path,
                        it.get("original_name",""),
                        it.get("source",""),
                        it.get("ts_str",""),
                    ])
            # Add CSV manifest
            csv_io = io.StringIO()
            writer = csv.writer(csv_io)
            writer.writerow(["ISSUE_ID","WARD","STATUS","SAVED_FILENAME","ORIGINAL_NAME","SOURCE","TIMESTAMP"])
            writer.writerows(rows)
            zf.writestr("after_photos_manifest.csv", csv_io.getvalue())

        zip_buf.seek(0)
        zip_name = datetime.now().strftime("after_photos_%Y-%m-%d_%H%M.zip")
        st.download_button(
            "⬇️ Download All After Photos (ZIP + manifest)",
            data=zip_buf,
            file_name=zip_name,
            mime="application/zip",
            use_container_width=True,
            help="Includes only photos from THIS session (restored counts do not include files)"
        )
    else:
        st.info("No After Photos captured/uploaded in this session yet.")
    # -------------------------------------------------------------------------------

    # Google Maps link for this batch (origin optional)
    if not batch_df.empty:
        waypoints = [(float(r['LATITUDE']), float(r['LONGITUDE'])) for _, r in batch_df.iterrows()]
        last = waypoints[-1]
        mids = waypoints[:-1] if len(waypoints) > 1 else []
        nav_url = google_maps_url(origin_lat, origin_lon, last[0], last[1], waypoints=mids)
        st.markdown(f"[🧭 Open route in Google Maps for this batch]({nav_url})")

    # Actions
    c1, c2, c3, c4 = st.columns(4)
    first_id = str(batch_df.iloc[0]['ISSUE ID']) if not batch_df.empty else None

    def _persist_counts_now():
        counts = {iid: len(items) for iid, items in st.session_state.uploaded_after_photos.items()}
        save_progress(st.session_state.dataset_id,
                      st.session_state.visited_ticket_ids,
                      st.session_state.skipped_ticket_ids,
                      counts)

    with c1:
        if st.button("✅ Mark first as Visited", key="btn_mark_first"):
            if first_id:
                st.session_state.visited_ticket_ids.add(first_id)
                _persist_counts_now()
                st.rerun()

    with c2:
        if st.button("⏭️ Skip first", key="btn_skip_first"):
            if first_id:
                st.session_state.skipped_ticket_ids.add(first_id)
                _persist_counts_now()
                st.rerun()

    with c3:
        if st.button("✅ Mark entire batch as Visited", key="btn_mark_batch"):
            for _id in batch_df['ISSUE ID'].astype(str).tolist():
                st.session_state.visited_ticket_ids.add(_id)
            _persist_counts_now()
            st.rerun()

    with c4:
        if st.button("➡️ Next batch", key="btn_next_batch"):
            st.session_state.batch_cursor = end if end < len(seq_df) else 0
            st.rerun()

# ---------------- Map (+ clickable image links / thumbnails) ----------------
st.subheader("Step 6: Map")

show_thumbs = st.checkbox("Show image thumbnails in popups (may slow the map)", value=False, key="cb_thumbs")

center_lat = float(df['LATITUDE'].mean())
center_lon = float(df['LONGITUDE'].mean())
m = folium.Map(location=[center_lat, center_lon], zoom_start=12)
mc = MarkerCluster(name="Tickets").add_to(m)

plot_df = (gdf_all.drop(columns=["geometry"]) if "geometry" in getattr(gdf_all, "columns", []) else gdf_all).copy()

# Color coding: visited = gray, skipped = purple, current batch = orange, first = green
batch_ids = set(batch_df['ISSUE ID'].astype(str).tolist()) if 'batch_df' in locals() and not batch_df.empty else set()
first_in_batch = str(batch_df.iloc[0]['ISSUE ID']) if 'batch_df' in locals() and not batch_df.empty else None

# Start marker if set
if origin_lat is not None and origin_lon is not None:
    folium.Marker([origin_lat, origin_lon], popup="Start here",
                  icon=folium.Icon(color="green", icon="flag")).add_to(m)

for _, row in plot_df.iterrows():
    rid = str(row['ISSUE ID'])
    lat, lon = float(row['LATITUDE']), float(row['LONGITUDE'])

    if rid == first_in_batch:
        color, size = 'green', 10
    elif rid in batch_ids:
        color, size = 'orange', 9
    elif rid in st.session_state.visited_ticket_ids:
        color, size = 'gray', 7
    elif rid in st.session_state.skipped_ticket_ids:
        color, size = 'purple', 7
    else:
        color, size = 'red', 7

    before_url = str(row.get('BEFORE PHOTO', '') or '').strip()
    ward       = escape(str(row.get('WARD', '') or ''))
    status     = escape(str(row.get('STATUS', '') or ''))
    cluster    = row.get('CLUSTER NUMBER', '')
    cluster    = int(cluster) if isinstance(cluster, (int, float)) and not pd.isna(cluster) else ''

    parts = [
        f"<b>Issue ID:</b> {escape(rid)}",
        f"<b>Status:</b> {status}",
        f"<b>Ward:</b> {ward}",
        f"<b>Lat, Lon:</b> {lat:.6f}, {lon:.6f}",
    ]
    if cluster != '':
        parts.insert(0, f"<b>Cluster:</b> {cluster}")

    if is_url(before_url):
        parts.append(f"<a href='{before_url}' target='_blank'>Before photo</a>")
        if show_thumbs:
            parts.append(f"<div><img src='{before_url}' style='max-width:220px;border:1px solid #ccc;border-radius:6px;'/></div>")

    # NOTE: Per your instruction, no After Photo link is shown anywhere.

    popup_html = "<div style='font-size:13px;line-height:1.25'>" + "<br>".join(parts) + "</div>"
    popup = folium.Popup(popup_html, max_width=280)

    folium.CircleMarker(
        location=[lat, lon],
        radius=size,
        color=color,
        fill=True,
        fill_color=color,
        fill_opacity=0.9,
        popup=popup
    ).add_to(mc)

m.save("Clustering_Application_Map.html")

# ---------------- Downloads ----------------
st.subheader("Step 7: Downloads")
if not clustered.empty:
    with open("Clustering_Application_Summary.xlsx", "rb") as f:
        st.download_button("Download Excel", f, file_name="Clustering_Application_Summary.xlsx", key="dl_excel")
with open("Clustering_Application_Map.html", "rb") as f:
    st.download_button("Download Map (HTML)", f, file_name="Clustering_Application_Map.html", key="dl_html")
